import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1'] #cell = sheet.cell(1, 1)
    # print(cell.value)

    # for row in range(1, sheet.max_row + 1):
    #   print(row)

    #for row in range(, sheet.max_row + 1):
    #    cell = sheet.cell(row, 3)
    #    print(cell.value)


    # Update Spreadsheet:
    for row in range(, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # Create a Chart
    chart = Barchart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    # Save changes in our spreadsheet:
    wb.save(filename)

